import io
import json
import os
import tempfile
from datetime import datetime
from itertools import product
from typing import Iterable, List

import pandas as pd
import plotly.graph_objects as go
import pyarrow as pa
import pyarrow.parquet as pq
from bson import ObjectId
from celery import states
from openpyxl import load_workbook
import psutil

from app.core.celery_app import celery_app
from app.core.config import settings
from app.core.minio_client import get_minio_client
from app.core.redis_client import get_sync_redis
from app.db.sync_mongo import get_sync_db


def _set_status(redis, viz_id: str, status: str, progress: int, message: str):
    pipe = redis.pipeline()
    name = f"visualization:{viz_id}:status"
    pipe.hset(name, "status", status)
    pipe.hset(name, "progress", progress)
    pipe.hset(name, "message", message)
    pipe.execute()


def _publish(viz_id: str, status: str, progress: int, message: str = ""):
    redis = get_sync_redis()
    payload = json.dumps(
        {"status": status, "progress": progress, "message": message or status}
    )
    redis.publish(f"visualization:{viz_id}:events", payload)
    _set_status(redis, viz_id, status, progress, message or status)

    # Mirror the latest task status in Mongo so API callers can still observe
    # progress even if Redis state is unavailable or evicted.
    db = get_sync_db()
    db.visualizations.update_one(
        {"_id": ObjectId(viz_id)},
        {
            "$set": {
                "status": status,
                "progress": progress,
                "message": message or status,
                "updated_at": datetime.utcnow(),
            }
        },
    )


def _iter_csv_frames(
    path: str,
    header_mode: str,
    custom_headers: list[str] | None,
    chunk_size: int,
    delim_whitespace: bool = False,
) -> Iterable[pd.DataFrame]:
    read_kwargs: dict = {
        "chunksize": chunk_size,
        "engine": "python" if delim_whitespace else "c",
    }
    if delim_whitespace:
        read_kwargs["delim_whitespace"] = True
    if header_mode == "none":
        read_kwargs["header"] = None
    elif header_mode == "custom":
        read_kwargs["header"] = None
        if custom_headers:
            read_kwargs["names"] = custom_headers
    else:
        read_kwargs["header"] = 0

    columns_cache: List[str] | None = None
    for chunk in pd.read_csv(path, **read_kwargs):
        if header_mode == "none" and not custom_headers:
            if columns_cache is None:
                columns_cache = [f"column_{i+1}" for i in range(len(chunk.columns))]
            chunk.columns = columns_cache
        yield chunk


def _iter_excel_frames(
    path: str, header_mode: str, custom_headers: list[str] | None, chunk_size: int
) -> Iterable[pd.DataFrame]:
    """Stream rows from an Excel worksheet without re-reading the file.

    pandas.read_excel() rereads from the start of the file on every call when
    combined with skiprows/nrows. That quickly becomes quadratic (and appears to
    "hang") on large spreadsheets. Using openpyxl in read-only mode lets us
    stream rows once and emit DataFrames in windowed batches.
    """

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)

        headers: List[str] | None = None
        if header_mode == "file":
            try:
                first_row = next(rows)
            except StopIteration:
                return
            headers = [
                str(cell) if cell is not None else f"column_{idx+1}"
                for idx, cell in enumerate(first_row)
            ]
        elif header_mode == "custom" and custom_headers:
            headers = list(custom_headers)

        chunk: list[list] = []
        for row in rows:
            chunk.append(list(row))
            if len(chunk) >= chunk_size:
                df = pd.DataFrame(chunk)
                chunk = []

                if headers:
                    width = len(headers)
                    if df.shape[1] < width:
                        df = df.reindex(columns=range(width))
                    df = df.iloc[:, :width]
                    df.columns = headers
                elif header_mode == "none" and not custom_headers:
                    df.columns = [f"column_{i+1}" for i in range(len(df.columns))]

                yield df

        if chunk:
            df = pd.DataFrame(chunk)
            if headers:
                width = len(headers)
                if df.shape[1] < width:
                    df = df.reindex(columns=range(width))
                df = df.iloc[:, :width]
                df.columns = headers
            elif header_mode == "none" and not custom_headers:
                df.columns = [f"column_{i+1}" for i in range(len(df.columns))]
            yield df
    finally:
        wb.close()


def _iter_frames(
    path: str, ext: str, header_mode: str, custom_headers: list[str] | None, chunk_size: int
) -> Iterable[pd.DataFrame]:
    ext = ext.lower()
    if ext in [".csv"]:
        yield from _iter_csv_frames(path, header_mode, custom_headers, chunk_size)
    elif ext in [".dat", ".txt"]:
        yield from _iter_csv_frames(
            path, header_mode, custom_headers, chunk_size, delim_whitespace=True
        )
    elif ext in [".xlsx", ".xlsm", ".xls"]:
        yield from _iter_excel_frames(path, header_mode, custom_headers, chunk_size)
    else:
        yield from _iter_csv_frames(path, header_mode, custom_headers, chunk_size)


def _download_object(minio, bucket: str, key: str, dest_path: str):
    response = minio.get_object(bucket, key)
    with open(dest_path, "wb") as fh:
        for data in response.stream(1 * 1024 * 1024):
            fh.write(data)


def _stat_size(minio, bucket: str, key: str) -> int:
    try:
        meta = minio.stat_object(bucket, key)
        return getattr(meta, "size", 0) or 0
    except Exception:
        return 0


def select_chunk_size(file_size_bytes: int, combo_count: int, base_chunk_size: int = 50000) -> int:
    """Choose a chunk size that balances RAM usage and throughput.

    The calculation leans on available memory, file size, and how many axis
    combinations will be expanded per chunk. It intentionally caps the
    resulting chunk size to avoid overwhelming the worker while still keeping
    throughput reasonable for smaller files.
    """

    mem = psutil.virtual_memory()
    safety_window = max(64 * 1024 * 1024, int(mem.available * 0.08))
    per_row_estimate = 256 * max(combo_count, 1)

    budget_rows = safety_window // per_row_estimate
    budget_rows = max(1000, min(200000, budget_rows))

    if file_size_bytes:
        target_chunks = max(1, min(32, file_size_bytes // (64 * 1024 * 1024)))
        rows_by_size = (file_size_bytes // max(target_chunks, 1)) // per_row_estimate
        rows_by_size = max(1000, min(200000, int(rows_by_size)))
        budget_rows = min(budget_rows, rows_by_size)

    if base_chunk_size:
        budget_rows = min(budget_rows, int(base_chunk_size))

    return int(max(1000, budget_rows))


def _figure_from_table(table: pa.Table):
    df = table.to_pandas()
    fig = go.Figure()
    z_present = "z" in df.columns
    for series_name, group in df.groupby("series"):
        if z_present and "z" in group.columns:
            fig.add_trace(
                go.Scatter3d(
                    x=group["x"],
                    y=group["y"],
                    z=group["z"],
                    mode="markers",
                    name=series_name,
                    marker={"size": 2},
                )
            )
        else:
            fig.add_trace(
                go.Scattergl(
                    x=group["x"],
                    y=group["y"],
                    mode="markers",
                    name=series_name,
                    marker={"size": 2},
                )
            )

    fig.update_layout(
        autosize=True,
        height=720,
        showlegend=True,
        margin={"t": 24, "r": 12, "l": 48, "b": 48},
    )
    return fig


def _numeric(series: pd.Series) -> pd.Series:
    """Best-effort numeric coercion with float dtype for Arrow parity."""

    coerced = pd.to_numeric(series, errors="coerce")
    if coerced.dtype != "float64":
        coerced = coerced.astype("float64")
    return coerced


@celery_app.task(bind=True, name=f"{settings.celery_task_prefix}.build_visualization")
def build_visualization(self, viz_id: str):
    redis = get_sync_redis()
    db = get_sync_db()
    minio = get_minio_client()

    doc = db.visualizations.find_one({"_id": ObjectId(viz_id)})
    if not doc:
        return

    bucket = settings.visualization_bucket
    if not minio.bucket_exists(bucket):
        minio.make_bucket(bucket)

    chunk_size = doc.get("chunk_size", 50000)
    chunk_prefix = doc.get("chunk_prefix")
    image_format = doc.get("image_format", "png")
    series = doc.get("series") or []

    _publish(viz_id, states.STARTED, 5, "Preparing visualization")
    db.visualizations.update_one(
        {"_id": ObjectId(viz_id)},
        {"$set": {"status": states.STARTED, "progress": 5, "updated_at": datetime.utcnow()}},
    )

    chunk_index = 0
    total_rows = 0
    trace_labels: list[dict] = []

    combined_writer: pq.ParquetWriter | None = None
    combined_path: str | None = None

    try:
        for series_idx, s in enumerate(series):
            job_id = s.get("job_id")
            job = db.ingestion_jobs.find_one({"_id": ObjectId(job_id)})
            if not job:
                raise ValueError(f"Ingestion job {job_id} missing")

            filename = (job.get("filename") or "")[0:256]
            ext = os.path.splitext(filename.lower())[-1]

            tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext if ext else "")
            os.close(tmp_fd)
            try:
                _publish(
                    viz_id,
                    states.STARTED,
                    max(5, int(10 + (series_idx / max(len(series), 1)) * 20)),
                    f"Downloading {job.get('filename', 'file')} from object store",
                )
                _download_object(
                    minio, settings.ingestion_bucket, job["storage_key"], tmp_path
                )

                combos = list(
                    product(
                        s.get("x_axes") or [],
                        s.get("y_axes") or [],
                        (s.get("z_axes") or [None]),
                    )
                )
                header_mode = job.get("header_mode", "file")
                custom_headers = job.get("custom_headers")

                combo_count = max(1, len(combos))
                file_size_bytes = _stat_size(
                    minio, settings.ingestion_bucket, job["storage_key"]
                )
                dynamic_chunk_size = select_chunk_size(
                    file_size_bytes, combo_count, base_chunk_size=chunk_size
                )
                if not combined_path:
                    handle = tempfile.NamedTemporaryFile(delete=False)
                    combined_path = handle.name
                    handle.close()

                for chunk in _iter_frames(
                    tmp_path, ext, header_mode, custom_headers, dynamic_chunk_size
                ):
                    output_tables: list[pa.Table] = []
                    for x_col, y_col, z_col in combos:
                        label = s.get("label") or job.get("filename") or "series"
                        trace_label = (
                            f"{label}: {x_col} vs {y_col}"
                            + (f" vs {z_col}" if z_col else "")
                        )
                        if not any(t.get("label") == trace_label for t in trace_labels):
                            trace_labels.append({"label": trace_label, "has_z": bool(z_col)})

                        payload = {
                            "x": _numeric(chunk[x_col]),
                            "y": _numeric(chunk[y_col]),
                            "series": trace_label,
                        }
                        if z_col:
                            payload["z"] = _numeric(chunk[z_col])
                        output_tables.append(
                            pa.Table.from_pandas(pd.DataFrame(payload), preserve_index=False)
                        )

                    if not output_tables:
                        continue
                    combined = (
                        pa.concat_tables(output_tables)
                        if len(output_tables) > 1
                        else output_tables[0]
                    )
                    total_rows += combined.num_rows

                    if combined_writer is None:
                        combined_writer = pq.ParquetWriter(combined_path, combined.schema)
                    combined_writer.write_table(combined)

                    tmp_out = tempfile.NamedTemporaryFile(delete=False)
                    tmp_out.close()
                    pq.write_table(combined, tmp_out.name)
                    object_key = f"{chunk_prefix}/chunk_{chunk_index}.parquet"
                    with open(tmp_out.name, "rb") as data:
                        stat = os.stat(tmp_out.name)
                        minio.put_object(
                            bucket_name=bucket,
                            object_name=object_key,
                            data=data,
                            length=stat.st_size,
                            content_type="application/octet-stream",
                        )
                    os.remove(tmp_out.name)

                    try:
                        fig = _figure_from_table(combined)
                        image_bytes = fig.to_image(
                            format=image_format,
                            engine="kaleido",
                            width=1280,
                            height=720,
                            scale=1,
                        )
                        img_obj = io.BytesIO(image_bytes)
                        image_key = (
                            f"{chunk_prefix}/images/chunk_{chunk_index}.{image_format}"
                        )
                        minio.put_object(
                            bucket_name=bucket,
                            object_name=image_key,
                            data=img_obj,
                            length=len(image_bytes),
                            content_type="image/png",
                        )
                    except Exception:
                        # Keep plot generation best-effort; data chunks remain available
                        pass
                    chunk_index += 1

                progress = int(30 + ((series_idx + 1) / max(len(series), 1)) * 50)
                _publish(
                    viz_id,
                    states.STARTED,
                    min(progress, 90),
                    f"Processed {series_idx + 1}/{len(series)} source files",
                )
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

        db.visualizations.update_one(
            {"_id": ObjectId(viz_id)},
            {
                "$set": {
                    "status": states.SUCCESS,
                    "progress": 100,
                    "chunk_count": chunk_index,
                    "rows_total": total_rows,
                    "trace_labels": trace_labels,
                    "updated_at": datetime.utcnow(),
                }
            },
        )
        if combined_writer:
            combined_writer.close()
        if combined_path and os.path.exists(combined_path):
            try:
                combined_table = pq.read_table(combined_path)
                fig = _figure_from_table(combined_table)
                final_image = fig.to_image(
                    format=image_format,
                    engine="kaleido",
                    width=1280,
                    height=720,
                    scale=1,
                )
                final_key = f"{chunk_prefix}/images/final.{image_format}"
                minio.put_object(
                    bucket_name=bucket,
                    object_name=final_key,
                    data=io.BytesIO(final_image),
                    length=len(final_image),
                    content_type="image/png",
                )
            finally:
                try:
                    os.remove(combined_path)
                except Exception:
                    pass
        _publish(viz_id, states.SUCCESS, 100, "Visualization ready")
    except Exception as exc:  # noqa: BLE001
        if combined_writer:
            try:
                combined_writer.close()
            except Exception:
                pass
        if combined_path and os.path.exists(combined_path):
            try:
                os.remove(combined_path)
            except Exception:
                pass
        _set_status(redis, viz_id, states.FAILURE, 100, str(exc))
        redis.publish(
            f"visualization:{viz_id}:events",
            json.dumps({"status": states.FAILURE, "progress": 100, "message": str(exc)}),
        )
        db.visualizations.update_one(
            {"_id": ObjectId(viz_id)},
            {
                "$set": {
                    "status": states.FAILURE,
                    "progress": 100,
                    "message": str(exc),
                    "updated_at": datetime.utcnow(),
                }
            },
        )
        raise
